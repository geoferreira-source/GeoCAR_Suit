# -*- coding: utf-8 -*-
"""excel_exporter.py — Gera relatório Excel com todas as bases incluindo SIGEF."""

import logging
from pathlib import Path
from datetime import datetime
log = logging.getLogger("GeoCAR_Suite")

COLUNAS = [
    ("codigo_car",                  46, "Código CAR"),
    ("uf",                           6, "UF"),
    ("cod_ibge",                    10, "Cód. IBGE"),
    ("municipio",                   22, "Município"),
    ("area_ha",                     14, "Área Total (ha)"),
    ("sobr_uc_federal_sim_nao",     10, "UC Federal"),
    ("sobr_uc_federal_ha",          14, "UC Federal (ha)"),
    ("sobr_uc_estadual_sim_nao",    10, "UC Estadual"),
    ("sobr_uc_estadual_ha",         14, "UC Estadual (ha)"),
    ("sobr_terra_indigena_sim_nao", 10, "Terra Indígena"),
    ("sobr_terra_indigena_ha",      14, "TI (ha)"),
    ("sobr_prodes_sim_nao",         10, "PRODES"),
    ("sobr_prodes_ha",              14, "PRODES Total (ha)"),
    ("prodes_anos_detalhe",         45, "PRODES por Ano"),
    ("sobr_assentamento_sim_nao",   12, "Assentamento"),
    ("sobr_assentamento_ha",        14, "Assentamento (ha)"),
    ("sobr_quilombo_sim_nao",       10, "Quilombo"),
    ("sobr_quilombo_ha",            14, "Quilombo (ha)"),
    ("sobr_embargo_ibama_sim_nao",  12, "Embargo IBAMA"),
    ("sobr_embargo_ibama_ha",       14, "Embargo (ha)"),
    ("sobr_sigef_sim_nao",          10, "SIGEF"),
    ("sobr_sigef_ha",               14, "SIGEF (ha)"),
    ("demonstrativo_pdf",           38, "Demonstrativo (PDF)"),
    ("status",                      10, "Status"),
    ("erro",                        40, "Erro / Observação"),
]

def exportar_excel(resultados: list, caminho_saida: str) -> tuple:
    if not resultados:
        return False, "Nenhum resultado para exportar"
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados CAR"

        def fill(cor): return PatternFill("solid", fgColor=cor)
        borda = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        centro = Alignment(horizontal="center", vertical="center")
        esq    = Alignment(horizontal="left",   vertical="center")

        # Título
        n = len(COLUNAS)
        ws.merge_cells(f"A1:{get_column_letter(n)}1")
        c = ws["A1"]
        c.value     = f"GeoCAR Suite — Relatório  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        c.font      = Font(bold=True, size=12, color="FFFFFF")
        c.fill      = fill("1B5E20")
        c.alignment = centro
        ws.row_dimensions[1].height = 22

        # Cabeçalho
        for ci, (campo, larg, titulo) in enumerate(COLUNAS, start=1):
            c = ws.cell(row=2, column=ci, value=titulo)
            c.font      = Font(bold=True, color="FFFFFF", size=10)
            c.fill      = fill("2E7D32")
            c.alignment = centro
            c.border    = borda
            ws.column_dimensions[get_column_letter(ci)].width = larg
        ws.row_dimensions[2].height = 20

        # Dados
        for ri, res in enumerate(resultados, start=3):
            tem_problema = (
                res.get("sobr_uc_federal_sim_nao") == "SIM" or
                res.get("sobr_terra_indigena_sim_nao") == "SIM" or
                res.get("sobr_embargo_ibama_sim_nao") == "SIM" or
                bool(res.get("erro"))
            )
            bg = fill("FFCDD2") if tem_problema else fill("E8F5E9" if ri%2 else "F5F5F5")

            for ci, (campo, _, _) in enumerate(COLUNAS, start=1):
                valor = res.get(campo, "")
                if campo == "area_ha" and isinstance(valor, (int, float)):
                    valor = round(valor, 2)
                c = ws.cell(row=ri, column=ci, value=valor)
                c.fill      = bg
                c.border    = borda
                c.alignment = centro if campo in ("uf","cod_ibge","area_ha","status",
                                                   "sobr_uc_federal_sim_nao","sobr_uc_estadual_sim_nao",
                                                   "sobr_terra_indigena_sim_nao","sobr_prodes_sim_nao",
                                                   "sobr_assentamento_sim_nao","sobr_quilombo_sim_nao",
                                                   "sobr_embargo_ibama_sim_nao","sobr_sigef_sim_nao") else esq
                c.font = Font(size=10)
                if campo == "status":
                    if str(valor).lower() == "ok":
                        c.font = Font(bold=True, color="1B5E20", size=10)
                    elif str(valor).lower() == "erro":
                        c.font = Font(bold=True, color="B71C1C", size=10)
                elif campo.endswith("_sim_nao"):
                    if str(valor) == "SIM":
                        c.font = Font(bold=True, color="B71C1C", size=10)
                        c.fill = fill("FFCDD2")
                    else:
                        c.font = Font(bold=True, color="1B5E20", size=10)
            ws.row_dimensions[ri].height = 16

        # Totais
        lt = len(resultados) + 3
        ws.merge_cells(f"A{lt}:{get_column_letter(n-2)}{lt}")
        ok_c  = sum(1 for r in resultados if not r.get("erro"))
        err_c = sum(1 for r in resultados if r.get("erro"))
        ws[f"A{lt}"].value     = f"Total: {len(resultados)}  |  ✓ Sucesso: {ok_c}  |  ✗ Erro: {err_c}"
        ws[f"A{lt}"].font      = Font(bold=True, size=10, color="FFFFFF")
        ws[f"A{lt}"].fill      = fill("1B5E20")
        ws[f"A{lt}"].alignment = centro

        ws.freeze_panes = "A3"
        Path(caminho_saida).parent.mkdir(parents=True, exist_ok=True)
        wb.save(caminho_saida)
        log.info(f"Excel: {caminho_saida}")
        return True, caminho_saida
    except ImportError:
        return False, "openpyxl não instalado"
    except Exception as e:
        return False, f"Erro ao gerar Excel: {e}"
